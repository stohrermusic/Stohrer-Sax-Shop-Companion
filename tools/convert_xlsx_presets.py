"""
Convert Music Center xlsx pad measurements to pad_presets.json format.
One-time conversion tool.
"""
import openpyxl
import json

wb = openpyxl.load_workbook('C:/Users/abadc/Downloads/Measures pads sax Music Center.xlsx')
ws = wb['SAXOS']

# Build grid
grid = {}
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            grid[(cell.row, cell.column)] = cell.value

max_row = ws.max_row

def fmt_size(val):
    """Format a numeric size cleanly (strip .0 from whole numbers)."""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val)

def is_section_header(row):
    """Check if row contains a section header like 'SAXO SOPRANO'."""
    for c in range(1, 25):
        val = grid.get((row, c))
        if isinstance(val, str) and 'SAXO' in val.upper():
            text = val.strip().upper()
            if 'SOPRANO' in text:
                return 'Soprano'
            elif 'ALTO' in text:
                return 'Alto'
            elif 'TENOR' in text:
                return 'Tenor'
            elif 'BARITONO' in text:
                return 'Bari'
    return None

def get_text_cells(row):
    """Get all text cells in a row as (col, value) pairs."""
    result = []
    for c in range(1, 25):
        val = grid.get((row, c))
        if isinstance(val, str) and val.strip():
            result.append((c, val.strip()))
    return result

def get_numeric_cells(row):
    """Get all numeric cells in a row as (col, value) pairs."""
    result = []
    for c in range(1, 25):
        val = grid.get((row, c))
        if isinstance(val, (int, float)):
            result.append((c, val))
    return result

def is_model_row(text_cells):
    """Check if a row of text cells looks like model names (multiple text entries)."""
    if len(text_cells) < 2:
        return False
    # Filter out brand-only headers
    names = [t[1] for t in text_cells]
    if len(names) == 1 and names[0].upper() in ('YAMAHA', 'SELMER'):
        return False
    return True

def is_brand_header(text_cells):
    """Check if row is a single brand name header."""
    if len(text_cells) == 1:
        name = text_cells[0][1].upper()
        return name in ('YAMAHA', 'SELMER')
    return False

def determine_brand(model_name, brand_context):
    """Figure out the brand for a model name."""
    name = model_name.upper()
    # Yamaha models
    if name.strip() == 'YAMAHA' or name.startswith('YSS') or name.startswith('YAS') or name.startswith('YTS'):
        return 'Yamaha'
    # Selmer models
    if any(s in name for s in ['MARK VI', 'MARK VII', 'SA 80', 'SERIE III', 'REFERENCE']):
        return 'Selmer'
    # Self-branded
    if any(s in name for s in ['JUPITER', 'YANAGISAWA', 'KEILWERTH', 'BUFFET', 'LUCETTE']):
        # Extract just the brand part
        for brand in ['JUPITER', 'YANAGISAWA', 'KEILWERTH', 'BUFFET', 'LUCETTE']:
            if brand in name:
                return brand.title()
        return model_name.strip()
    if 'PRESTIGE' in name:
        return 'Buffet'
    # Fall back to context
    if brand_context:
        return brand_context
    return ''

def clean_model_name(model_name, brand):
    """Clean up model name for the preset, removing redundant brand prefix."""
    name = model_name.strip()
    upper = name.upper()

    # For self-branded entries like "JUPITER" or "YANAGISAWA" alone,
    # there's no separate model - just use the brand
    if upper in ('JUPITER', 'YANAGISAWA', 'BUFFET', 'LUCETTE', 'YAMAHA'):
        return ''

    # Clean up Yamaha model names
    if upper.startswith('YSS') or upper.startswith('YAS') or upper.startswith('YTS'):
        # "YSS 62" -> "YSS-62", "YAS23-YAS275-YAS475" -> "YAS-23/275/475"
        # "YTS275 - YTS52- YTS62" -> "YTS-275/52/62"
        # "YAS875 CUSTOM" -> "YAS-875 Custom"
        # "YAS82 Z" -> "YAS-82Z"
        parts = upper.replace(' - ', '-').replace('- ', '-').replace(' -', '-').split('-')
        parts = [p.strip() for p in parts if p.strip()]
        if len(parts) > 1:
            # Multiple models like YAS23, YAS275, YAS475
            prefix = ''
            numbers = []
            for p in parts:
                for i, ch in enumerate(p):
                    if ch.isdigit():
                        pref = p[:i]
                        num = p[i:]
                        if not prefix:
                            prefix = pref
                        numbers.append(num)
                        break
            if prefix and numbers:
                return f"{prefix}-{'/'.join(numbers)}"
        else:
            # Single model like "YSS 62", "YAS875 CUSTOM", "YAS82 Z"
            p = parts[0]
            # Find where digits start
            for i, ch in enumerate(p):
                if ch.isdigit():
                    prefix = p[:i]
                    rest = p[i:]
                    # Handle suffix like "CUSTOM" or "Z" after the number
                    # rest might be "875CUSTOM" or "82Z" or "875 CUSTOM"
                    # Use the original (not uppercased) for nice casing
                    orig = name.strip()
                    for j, och in enumerate(orig):
                        if och.isdigit():
                            orig_prefix = orig[:j].strip().upper()
                            orig_rest = orig[j:]
                            # Clean up: "875 CUSTOM" -> "875 Custom", "82 Z" -> "82Z"
                            orig_rest = orig_rest.strip()
                            # Title-case any alpha suffix
                            result_parts = []
                            for word in orig_rest.split():
                                if word.isdigit():
                                    result_parts.append(word)
                                elif word.isalpha():
                                    result_parts.append(word.title())
                                else:
                                    result_parts.append(word)
                            return f"{orig_prefix}-{' '.join(result_parts)}"
                    return f"{prefix}-{rest}"
            return p

    # Keilwerth - strip "KEILWERTH" prefix and clean up
    if 'KEILWERTH' in upper:
        cleaned = upper.replace('KEILWERTH', '').replace('- SERIES', '').replace('-SERIES', '').strip()
        if cleaned:
            return cleaned  # "SX", "ST"
        return name

    # Prestige
    if 'PRESTIGE' in upper:
        return name.title()

    return name

def build_preset_name(instrument, brand, model_clean):
    """Build a clean preset name."""
    if model_clean:
        return f"{instrument} - {brand} {model_clean}"
    else:
        return f"{instrument} - {brand}"


# ---- Main parsing ----

presets = {}
current_instrument = None
brand_context = None
active_models = []  # list of (start_col, brand, model_name, clean_name)
model_pads = {}     # key = preset_name, value = list of (qty, size)

def flush_models():
    """Save accumulated pad data for active models."""
    global model_pads
    for start_col, brand, raw_name, preset_name in active_models:
        if preset_name in model_pads and model_pads[preset_name]:
            raw_pads = model_pads[preset_name]
            # Consolidate: sum quantities for each unique size
            size_counts = {}
            for qty, size in raw_pads:
                size_counts[size] = size_counts.get(size, 0) + qty
            # Sort largest-first
            sorted_sizes = sorted(size_counts.keys(), reverse=True)
            parts = []
            for size in sorted_sizes:
                qty = size_counts[size]
                size_str = fmt_size(size)
                if qty > 1:
                    parts.append(f"{size_str}x{qty}")
                else:
                    parts.append(size_str)
            pad_string = ", ".join(parts)
            presets[preset_name] = {
                "pads": pad_string,
                "notes": "via Oscar @ Ribera Luthier"
            }

for r in range(1, max_row + 1):
    # Check for section header
    section = is_section_header(r)
    if section:
        flush_models()
        current_instrument = section
        brand_context = None
        active_models = []
        model_pads = {}
        continue

    if not current_instrument:
        continue

    text_cells = get_text_cells(r)
    num_cells = get_numeric_cells(r)

    # Brand header?
    if text_cells and is_brand_header(text_cells):
        brand_context = text_cells[0][1].strip().title()
        continue

    # Model name row?
    if text_cells and is_model_row(text_cells) and len(num_cells) == 0:
        # Flush previous model group
        flush_models()
        active_models = []
        model_pads = {}

        for col, name in text_cells:
            brand = determine_brand(name, brand_context)
            model_clean = clean_model_name(name, brand)
            preset_name = build_preset_name(current_instrument, brand, model_clean)
            active_models.append((col, brand, name, preset_name))
            model_pads[preset_name] = []
        continue

    # Data row - extract pad data for each active model
    if num_cells and active_models:
        for start_col, brand, raw_name, preset_name in active_models:
            qty_val = grid.get((r, start_col))
            size_val = grid.get((r, start_col + 1))
            if isinstance(qty_val, (int, float)) and isinstance(size_val, (int, float)):
                model_pads[preset_name].append((int(qty_val), size_val))

# Flush final group
flush_models()

# Build library structure
library = {"Music Center": presets}

# Output
output_path = 'C:/Users/abadc/Downloads/music_center_pad_presets.json'
with open(output_path, 'w') as f:
    json.dump(library, f, indent=2)

print(f"Generated {len(presets)} presets:")
for name, data in presets.items():
    pads = data['pads']
    # Count total pads
    total = 0
    for part in pads.split(', '):
        if 'x' in part:
            total += int(part.split('x')[1])
        else:
            total += 1
    print(f"  {name}: {total} pads — {pads}")

print(f"\nSaved to: {output_path}")
